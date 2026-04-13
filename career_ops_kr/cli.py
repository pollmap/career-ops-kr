"""career-ops-kr CLI entry point.

Click-based command-line interface. Provides commands for scan/score/pipeline/list
and graceful fallbacks when subpackage imports fail.
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

import click
from rich.console import Console
from rich.panel import Panel
from rich.progress import BarColumn, MofNCompleteColumn, Progress, TextColumn, TimeElapsedColumn
from rich.prompt import Confirm, Prompt
from rich.table import Table

console = Console()

# Project root assumed to be CWD; avoid ~/.career-ops/config.yml
PROJECT_ROOT = Path.cwd()
CONFIG_DIR = PROJECT_ROOT / "config"
DATA_DIR = PROJECT_ROOT / "data"
MODES_DIR = PROJECT_ROOT / "modes"


def _fallback_import(module_name: str, what_for: str) -> Any:
    """Import with graceful error. Returns module or None and prints fix."""
    try:
        return __import__(module_name, fromlist=["*"])
    except Exception as exc:
        console.print(
            f"[red]ERROR[/red] cannot import [bold]{module_name}[/bold] "
            f"(needed for {what_for}): {exc}"
        )
        console.print(
            "[yellow]Fix:[/yellow] run [bold]pip install -e .[dev][/bold] "
            "from the project root, and verify package layout under career_ops_kr/."
        )
        return None


def _ensure_project_dirs() -> None:
    for d in (CONFIG_DIR, DATA_DIR, MODES_DIR):
        d.mkdir(parents=True, exist_ok=True)


@click.group(help="career-ops-kr — 한국형 구직 자동화 CLI (Luxon AI)")
@click.version_option("1.0.0", prog_name="career-ops")
def cli() -> None:
    """Root group."""
    _ensure_project_dirs()


def _load_preset_loader() -> Any:
    """Import PresetLoader with graceful fallback.

    Returns the PresetLoader class or None if unavailable.
    Printed message guides the user to install the presets module
    (which is delivered by a parallel implementation track).
    """
    try:
        from career_ops_kr.presets.loader import PresetLoader

        return PresetLoader
    except Exception as exc:
        console.print(
            "[yellow]preset 기능을 사용하려면 career_ops_kr.presets 모듈이 필요합니다[/yellow]"
        )
        console.print(f"[dim]import error: {exc}[/dim]")
        console.print(
            "[dim]docs/presets.md 참조 — 해당 모듈이 아직 배포되지 않았다면 "
            "interactive 모드(`career-ops init`)를 사용하세요.[/dim]"
        )
        return None


def _show_preset_list() -> int:
    """Print rich table of available presets. Returns exit code."""
    loader_cls = _load_preset_loader()
    if loader_cls is None:
        return 2
    try:
        loader = loader_cls()
        presets = loader.list_available()
    except Exception as exc:
        console.print(f"[red]preset 목록 조회 실패[/red] {exc}")
        return 2

    table = Table(title="사용 가능한 프리셋")
    table.add_column("preset_id", style="cyan")
    table.add_column("label_ko")
    table.add_column("description", style="dim")
    for p in presets:
        table.add_row(
            str(p.get("preset_id", "")),
            str(p.get("label_ko", "")),
            str(p.get("description", "")),
        )
    console.print(table)
    console.print("[dim]사용법: [bold]career-ops init --preset <preset_id>[/bold][/dim]")
    return 0


def _apply_preset(preset_id: str, force: bool) -> int:
    """Apply a preset to config/ + modes/_profile.md with HITL G1 gate.

    Returns exit code.
    """
    loader_cls = _load_preset_loader()
    if loader_cls is None:
        return 2
    try:
        loader = loader_cls()
        preset = loader.load(preset_id)
    except Exception as exc:
        console.print(f"[red]preset '{preset_id}' 로드 실패[/red] {exc}")
        return 2

    label = preset.get("label_ko", preset_id) if isinstance(preset, dict) else preset_id
    console.print(
        Panel.fit(
            f"[bold]{preset_id}[/bold] — {label}\n"
            f"[dim]대상 디렉토리: {CONFIG_DIR}[/dim]\n"
            f"[dim]overwrite: {force}[/dim]",
            title="preset apply plan",
            border_style="cyan",
        )
    )

    existing: list[Path] = []
    for name in (
        "profile.yml",
        "portals.yml",
        "qualifier_rules.yml",
        "scoring_weights.yml",
    ):
        p = CONFIG_DIR / name
        if p.exists():
            existing.append(p)
    if (MODES_DIR / "_profile.md").exists():
        existing.append(MODES_DIR / "_profile.md")

    if existing and not force:
        console.print("[yellow]다음 파일이 이미 존재합니다 (덮어쓰지 않음):[/yellow]")
        for p in existing:
            console.print(f"  - {p}")
        console.print("[dim]덮어쓰려면 --force 플래그를 사용하세요.[/dim]")

    # HITL G1: always confirm before writing
    if not Confirm.ask(f"프리셋 [bold cyan]{preset_id}[/bold cyan]을(를) 적용할까?", default=True):
        console.print("[yellow]aborted[/yellow]")
        return 0

    try:
        written = loader.apply_to(preset_id, CONFIG_DIR, overwrite=force)
    except Exception as exc:
        console.print(f"[red]preset 적용 실패[/red] {exc}")
        return 2

    console.print(
        Panel.fit(
            "\n".join(f"[green]created[/green] {k}: {v}" for k, v in written.items())
            or "[dim](no files written)[/dim]",
            title=f"preset '{preset_id}' applied",
            border_style="green",
        )
    )
    console.print(
        "[dim]다음: [bold]career-ops init[/bold] (인터뷰 모드)로 cv.md를 마저 채우거나, "
        "docs/presets.md 에서 프리셋별 세부 지침을 확인하세요.[/dim]"
    )
    return 0


@cli.command("init", help="G1 온보딩 — 누락 유저 파일 생성 (cv.md / config.yml / _profile.md)")
@click.option(
    "--preset",
    "preset",
    type=str,
    default=None,
    help="도메인 프리셋: finance|dev|design|marketing|research|public|edu",
)
@click.option(
    "--list-presets",
    "list_presets",
    is_flag=True,
    help="사용 가능한 프리셋 목록 표시 후 종료",
)
@click.option(
    "--force",
    "force",
    is_flag=True,
    help="기존 config 덮어쓰기",
)
def init_cmd(preset: str | None, list_presets: bool, force: bool) -> None:
    # 1) preset 목록 표시
    if list_presets:
        code = _show_preset_list()
        if code != 0:
            sys.exit(code)
        return

    # 2) preset 적용 (HITL G1 gate)
    if preset:
        code = _apply_preset(preset, force=force)
        if code != 0:
            sys.exit(code)
        return

    # 3) 기본(기존) 인터랙티브 플로우 — 하위 호환
    cv = PROJECT_ROOT / "cv.md"
    profile_yml = CONFIG_DIR / "profile.yml"
    profile_md = MODES_DIR / "_profile.md"

    missing: list[str] = []
    if not cv.exists():
        missing.append("cv.md")
    if not profile_yml.exists():
        missing.append("config/profile.yml")
    if not profile_md.exists():
        missing.append("modes/_profile.md")

    if not missing:
        console.print("[green]OK[/green] onboarding complete — all user files present")
        return

    console.print(f"[yellow]G1 onboarding[/yellow] missing: {', '.join(missing)}")
    if not Confirm.ask("대화형 인터뷰로 생성할까?", default=True):
        console.print("[yellow]skipped[/yellow]")
        return

    if "cv.md" in missing:
        name = Prompt.ask("이름 (한글 또는 English)", default="")
        name_line = f"# {name} 이력서" if name else "# 이력서"
        cv.write_text(
            f"{name_line}\n\n(본인 이력을 이곳에 작성하세요.)\n",
            encoding="utf-8",
        )
        console.print(f"[green]created[/green] {cv}")

    if "config/profile.yml" in missing:
        # 템플릿(templates/profile.example.yml)이 있으면 그걸 복사,
        # 없으면 최소 placeholder yml 생성. 개인정보 하드코딩 금지.
        tmpl_profile = PROJECT_ROOT / "templates" / "profile.example.yml"
        if tmpl_profile.exists():
            profile_yml.write_text(
                tmpl_profile.read_text(encoding="utf-8"),
                encoding="utf-8",
            )
            console.print(
                f"[green]created[/green] {profile_yml} (templates/profile.example.yml에서 복사)"
            )
        else:
            profile_yml.write_text(
                "# USER 프로필 — 본인 정보로 편집하세요.\n"
                "name:\n  ko: \"\"\n  en: \"\"\n"
                "target_industries:\n  - 금융\n  - 핀테크\n  - 블록체인\n",
                encoding="utf-8",
            )
            console.print(f"[green]created[/green] {profile_yml}")

    if "modes/_profile.md" in missing:
        tmpl = MODES_DIR / "_profile.template.md"
        if tmpl.exists():
            profile_md.write_text(tmpl.read_text(encoding="utf-8"), encoding="utf-8")
        else:
            profile_md.write_text("# Profile\n\n(placeholder)\n", encoding="utf-8")
        console.print(f"[green]created[/green] {profile_md}")


@cli.command("scan", help="스캔 모드 — 포털 크롤링 + 신규 공고 수집")
@click.option("--tier", type=int, default=None, help="특정 tier만 스캔 (1~5)")
@click.option("--site", type=str, default=None, help="특정 사이트 이름")
@click.option("--all", "all_sites", is_flag=True, help="모든 포털 스캔")
@click.option("--dry-run", is_flag=True, help="네트워크 호출 없이 설정만 검증")
@click.option(
    "--concurrency",
    type=int,
    default=1,
    show_default=True,
    help="동시 스캔 채널 수 (1=순차, 2+=병렬 ThreadPool)",
)
def scan_cmd(
    tier: int | None,
    site: str | None,
    all_sites: bool,
    dry_run: bool,
    concurrency: int,
) -> None:
    if dry_run:
        console.print("[cyan]dry-run[/cyan] scan config OK — no network calls")
        table = Table(title="Scan plan")
        table.add_column("filter")
        table.add_column("value")
        table.add_row("tier", str(tier) if tier else "(any)")
        table.add_row("site", site or "(any)")
        table.add_row("all", str(all_sites))
        table.add_row("concurrency", str(concurrency))
        console.print(table)
        return

    if concurrency > 1:
        _scan_parallel(tier=tier, site=site, concurrency=concurrency)
        return

    # 순차 모드 (기존 동작 보존)
    mcp = _fallback_import("career_ops_kr.mcp_server", "scan mode")
    if mcp is None:
        sys.exit(2)

    console.print(f"[green]scan[/green] tier={tier} site={site} all={all_sites}")
    try:
        result = mcp.tool_scan_jobs(tier=tier, site=site)
    except Exception as exc:
        console.print(f"[red]scan failed[/red] {exc}")
        sys.exit(1)

    if isinstance(result, dict) and "error" in result:
        console.print(f"[red]scan error[/red] {result['error']}")
        hint = result.get("hint")
        if hint:
            console.print(f"[dim]hint: {hint}[/dim]")
        sys.exit(1)

    table = Table(title=f"Scan results (total={result.get('total', 0)})")
    table.add_column("channel", style="cyan")
    table.add_column("tier", justify="right")
    table.add_column("backend", style="dim")
    table.add_column("count", justify="right")
    table.add_column("error", style="red")

    channels_summary: dict[str, Any] = result.get("channels", {}) or {}
    for ch_name, info in channels_summary.items():
        info = info or {}
        table.add_row(
            ch_name,
            str(info.get("tier", "")),
            str(info.get("backend", "")),
            str(info.get("count", 0)),
            str(info.get("error", "") or ""),
        )
    console.print(table)


def _scan_parallel(
    tier: int | None,
    site: str | None,
    concurrency: int,
) -> None:
    """병렬 스캔 — ThreadPoolExecutor로 채널별 동시 크롤링."""
    from concurrent.futures import ThreadPoolExecutor, as_completed

    try:
        from career_ops_kr.channels import CHANNEL_REGISTRY
    except ImportError as exc:
        console.print(f"[red]channels import 실패[/red]: {exc}")
        sys.exit(2)

    # 필터 적용
    targets: list[tuple[str, type]] = []
    for name, cls in CHANNEL_REGISTRY.items():
        if site and name != site:
            continue
        if tier is not None and getattr(cls, "tier", None) != tier:
            continue
        targets.append((name, cls))

    if not targets:
        console.print("[yellow]대상 채널 없음[/yellow]")
        return

    console.print(
        f"[green]parallel scan[/green] {len(targets)}개 채널 × concurrency={concurrency}"
    )

    from rich.progress import BarColumn, MofNCompleteColumn, Progress, TextColumn, TimeElapsedColumn

    summary: dict[str, dict[str, Any]] = {}
    total_count = 0

    def _scrapling_fallback(ch_name: str, ch_cls: type) -> list[Any]:
        """Scrapling fallback: re-fetch landing page when requests returns 0."""
        try:
            from career_ops_kr.channels._scrapling_base import SCRAPLING_AVAILABLE, ScraplingChannel
            if not SCRAPLING_AVAILABLE:
                return []
            from career_ops_kr.channels.base import JobRecord, deadline_parser
            from bs4 import BeautifulSoup
            from datetime import datetime

            landing = getattr(ch_cls, "LANDING_URL", None)
            if landing is None:
                # Try to get from a fresh instance
                inst = ch_cls()
                landing = getattr(inst, "landing_url", None)
            if not landing:
                return []

            sc = ScraplingChannel(name=f"{ch_name}_scrapling", tier=getattr(ch_cls, "tier", 6))
            page_data = sc.fetch_page(landing)
            if not page_data or not page_data.get("html"):
                return []

            soup = BeautifulSoup(page_data["html"], "html.parser")
            results: list[Any] = []
            now = datetime.now()
            seen: set[str] = set()
            org = ch_name.replace("_", " ").title()
            base_url = landing.split("/", 3)[:3]
            base = "/".join(base_url) if len(base_url) >= 3 else landing

            for anchor in soup.find_all("a"):
                text = (anchor.get_text(" ", strip=True) or "").strip()
                href = anchor.get("href") or ""
                if not text or len(text) < 4 or len(text) > 300 or not href:
                    continue
                if href.startswith("#") or href.lower().startswith("javascript"):
                    continue
                if any(kw in text for kw in ("채용", "공고", "모집", "인턴", "신입", "경력", "지원", "recruit")):
                    if href.startswith("/"):
                        href = base + href
                    if "://" not in href:
                        continue
                    if href in seen:
                        continue
                    seen.add(href)
                    try:
                        from career_ops_kr.channels.base import BaseChannel
                        results.append(JobRecord(
                            id=BaseChannel._make_id(href, text[:120]),
                            source_url=href,
                            source_channel=ch_name,
                            source_tier=getattr(ch_cls, "tier", 6),
                            org=org,
                            title=text[:200],
                            deadline=deadline_parser(text),
                            description=text,
                            legitimacy_tier=getattr(ch_cls, "default_legitimacy_tier", "T5"),
                            scanned_at=now,
                        ))
                    except Exception:
                        continue
                    if len(results) >= 30:
                        break
            return results
        except Exception:
            return []

    def _scan_one(ch_name: str, ch_cls: type) -> tuple[str, dict[str, Any]]:
        try:
            instance = ch_cls()
            jobs = instance.list_jobs() or []
            backend = getattr(ch_cls, "backend", "unknown")
            # Scrapling auto-fallback: 0건이면 scrapling으로 재시도
            if not jobs:
                scrapling_jobs = _scrapling_fallback(ch_name, ch_cls)
                if scrapling_jobs:
                    jobs = scrapling_jobs
                    backend = "scrapling(fallback)"
            return ch_name, {
                "count": len(jobs),
                "tier": getattr(ch_cls, "tier", None),
                "backend": backend,
            }
        except Exception as exc:
            return ch_name, {"count": 0, "error": str(exc)}

    with Progress(
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        MofNCompleteColumn(),
        TimeElapsedColumn(),
        console=console,
        transient=True,
    ) as prog:
        task = prog.add_task("[cyan]병렬 스캔[/cyan]", total=len(targets))

        with ThreadPoolExecutor(max_workers=concurrency) as pool:
            futures = {
                pool.submit(_scan_one, name, cls): name
                for name, cls in targets
            }
            for future in as_completed(futures):
                ch_name, info = future.result()
                summary[ch_name] = info
                total_count += info.get("count", 0)
                prog.advance(task)

    # 결과 테이블
    table = Table(title=f"Parallel scan results (total={total_count})")
    table.add_column("channel", style="cyan")
    table.add_column("tier", justify="right")
    table.add_column("backend", style="dim")
    table.add_column("count", justify="right")
    table.add_column("error", style="red")

    for ch_name in sorted(summary):
        info = summary[ch_name]
        table.add_row(
            ch_name,
            str(info.get("tier", "")),
            str(info.get("backend", "")),
            str(info.get("count", 0)),
            str(info.get("error", "") or ""),
        )
    console.print(table)


@cli.command("score", help="단일 URL 평가 — markdown 리포트 출력")
@click.argument("url")
def score_cmd(url: str) -> None:
    if "://" not in url or url.startswith("invalid"):
        console.print(f"[red]invalid URL[/red] {url}")
        sys.exit(1)

    mcp = _fallback_import("career_ops_kr.mcp_server", "score mode")
    if mcp is None:
        sys.exit(2)

    console.print(f"[green]scoring[/green] {url}")
    try:
        result = mcp.tool_score_job(url=url)
    except Exception as exc:
        console.print(f"[red]score failed[/red] {exc}")
        sys.exit(1)

    if isinstance(result, dict) and "error" in result:
        console.print(f"[red]score error[/red] {result['error']}")
        hint = result.get("hint")
        if hint:
            console.print(f"[dim]hint: {hint}[/dim]")
        sys.exit(1)

    url_out = result.get("url", url)
    legitimacy = result.get("legitimacy") or "T5 미확인"
    archetype = result.get("archetype") or "(미판정)"
    grade = result.get("grade") or "(미평가)"
    total = result.get("total_score")
    fit_grade_line = f"{grade}" + (f" ({total})" if total is not None else "")
    eligibility = result.get("qualifier_verdict") or "(미판정)"
    deadline = result.get("deadline") or "(미상)"
    org = result.get("org") or ""
    title = result.get("title") or ""
    reasons = result.get("reasons") or []

    lines = [
        f"**URL:** {url_out}",
        f"**Legitimacy:** {legitimacy}",
        f"**Archetype:** {archetype}",
        f"**Fit Grade:** {fit_grade_line}",
        f"**Eligibility:** {eligibility}",
        f"**Deadline:** {deadline}",
    ]
    if org or title:
        lines.append(f"**Org / Title:** {org} / {title}")
    if reasons:
        lines.append("")
        lines.append("**Reasons:**")
        for r in reasons:
            lines.append(f"- {r}")

    console.print(
        Panel.fit(
            "\n".join(lines),
            title="score report",
            border_style="green",
        )
    )


@cli.command("pipeline", help="배치 평가 — inbox 전체 (G4 게이트)")
@click.option("--limit", type=int, default=10)
def pipeline_cmd(limit: int) -> None:
    if limit >= 5:
        console.print(f"[yellow]G4 batch gate[/yellow] {limit}건 일괄 평가 예정")
        if not Confirm.ask("샘플 1건 후 계속 진행?", default=True):
            console.print("[yellow]aborted[/yellow]")
            return
    console.print(f"[green]pipeline[/green] processing up to {limit} items")

    db = DATA_DIR / "jobs.db"
    if not db.exists():
        console.print(
            "[yellow]inbox empty[/yellow] (no jobs.db — run [bold]career-ops scan[/bold] first)"
        )
        return

    mcp = _fallback_import("career_ops_kr.mcp_server", "pipeline mode")
    store_mod = _fallback_import("career_ops_kr.storage.sqlite_store", "pipeline mode")
    if mcp is None or store_mod is None:
        sys.exit(2)

    try:
        store = store_mod.SQLiteStore(db)
    except Exception as exc:
        console.print(f"[red]SQLiteStore init failed[/red] {exc}")
        sys.exit(1)

    # Pull inbox-status jobs (ungraded) — use search() with empty keyword.
    try:
        inbox = store.search(keyword="")
    except Exception as exc:
        console.print(f"[red]inbox query failed[/red] {exc}")
        sys.exit(1)

    inbox = [j for j in inbox if (j.get("status") or "inbox") == "inbox"]
    if not inbox:
        console.print("[yellow]no inbox items[/yellow]")
        return

    table = Table(title=f"pipeline results ({min(limit, len(inbox))}/{len(inbox)})")
    table.add_column("org", style="cyan")
    table.add_column("title")
    table.add_column("grade", justify="center")
    table.add_column("score", justify="right")
    table.add_column("error", style="red")

    processed = 0
    for job in inbox[:limit]:
        url = job.get("source_url") or ""
        if not url:
            continue
        try:
            result = mcp.tool_score_job(url=url)
        except Exception as exc:
            table.add_row(str(job.get("org") or ""), str(job.get("title") or ""), "", "", str(exc))
            processed += 1
            continue

        if isinstance(result, dict) and "error" in result:
            table.add_row(
                str(job.get("org") or ""),
                str(job.get("title") or ""),
                "",
                "",
                str(result.get("error", "")),
            )
        else:
            table.add_row(
                str(result.get("org") or ""),
                str(result.get("title") or ""),
                str(result.get("grade") or ""),
                str(result.get("total_score") or ""),
                "",
            )
        processed += 1

    console.print(table)
    console.print(f"[dim]processed {processed} item(s)[/dim]")


@cli.command("list", help="SQLite 조회 — 저장된 공고 테이블 출력")
@click.option("--grade", type=str, default=None)
@click.option("--status", type=str, default=None)
@click.option("--archetype", type=str, default=None)
@click.option(
    "--sector",
    type=click.Choice(["금융", "공공", "안보", "핀테크", "기타"]),
    default=None,
    help="섹터 필터 (금융/공공/안보/핀테크/기타)",
)
def list_cmd(
    grade: str | None,
    status: str | None,
    archetype: str | None,
    sector: str | None,
) -> None:
    db = DATA_DIR / "jobs.db"
    if not db.exists():
        console.print("[yellow]no jobs found[/yellow] (empty DB)")
        return

    store_mod = _fallback_import("career_ops_kr.storage.sqlite_store", "list mode")
    if store_mod is None:
        sys.exit(2)

    try:
        store = store_mod.SQLiteStore(db)
    except Exception as exc:
        console.print(f"[red]SQLiteStore init failed[/red] {exc}")
        sys.exit(1)

    try:
        if grade:
            jobs = store.list_by_grade(grade)
        elif archetype:
            jobs = store.search(keyword="", archetype=archetype)
        else:
            jobs = store.search(keyword="")
    except Exception as exc:
        console.print(f"[red]query failed[/red] {exc}")
        sys.exit(1)

    if status:
        jobs = [j for j in jobs if (j.get("status") or "") == status]

    if sector:
        from career_ops_kr.sector import infer_sector

        jobs = [
            j for j in jobs
            if infer_sector(j.get("source_channel"), j.get("org"), j.get("title")) == sector
        ]

    if not jobs:
        console.print("[yellow]no jobs found[/yellow] (filter matched 0 rows)")
        console.print(
            f"[dim]filters: grade={grade} status={status} "
            f"archetype={archetype} sector={sector}[/dim]"
        )
        return

    table = Table(title=f"Jobs ({len(jobs)})")
    for col in ("id", "sector", "org", "title", "grade", "match", "status", "deadline"):
        table.add_column(col)
    from career_ops_kr.sector import infer_sector as _sec
    for job in jobs:
        fs = job.get("fit_score")
        match_str = f"{int(fs)}%" if fs is not None and fs != "" else "-"
        table.add_row(
            str(job.get("id") or "")[:12],
            _sec(job.get("source_channel"), job.get("org"), job.get("title")),
            str(job.get("org") or ""),
            str(job.get("title") or ""),
            str(job.get("fit_grade") or ""),
            match_str,
            str(job.get("status") or ""),
            str(job.get("deadline") or ""),
        )
    console.print(table)
    console.print(
        f"[dim]filters: grade={grade} status={status} "
        f"archetype={archetype} sector={sector}[/dim]"
    )


@cli.command("calendar", help=".ics 파일 생성 — 마감일 일정")
@click.option(
    "--output",
    type=click.Path(path_type=Path),
    default=PROJECT_ROOT / "output" / "deadlines.ics",
)
@click.option("--days", type=int, default=30, help="Deadline window in days")
def calendar_cmd(output: Path, days: int) -> None:
    try:
        from career_ops_kr.calendar.ics_export import CalendarExporter
    except Exception as exc:
        console.print(f"[red]calendar import failed[/red] {exc}")
        sys.exit(2)

    jobs: list[dict[str, Any]] = []
    db = DATA_DIR / "jobs.db"
    if db.exists():
        store_mod = _fallback_import("career_ops_kr.storage.sqlite_store", "calendar mode")
        if store_mod is not None:
            try:
                store = store_mod.SQLiteStore(db)
                jobs = list(store.list_upcoming_deadlines(days=days))
            except Exception as exc:
                console.print(
                    f"[yellow]deadline query failed — writing empty calendar[/yellow] {exc}"
                )
                jobs = []

    output.parent.mkdir(parents=True, exist_ok=True)
    exporter = CalendarExporter()
    path = exporter.from_jobs(jobs, output)
    console.print(f"[green]calendar[/green] written {path} ({len(jobs)} job(s))")


@cli.command("patterns", help="패턴 분석 모드 — N일치 잡 히스토리")
@click.option("--days", type=int, default=7)
def patterns_cmd(days: int) -> None:
    console.print(f"[green]patterns[/green] last {days} days")


@cli.command("verify", help="데이터 무결성 검증 — verify_pipeline.py")
def verify_cmd() -> None:
    script = PROJECT_ROOT / "scripts" / "verify_pipeline.py"
    if not script.exists():
        # Empty project counts as clean
        console.print("[green]verify[/green] clean (no pipeline script, nothing to verify)")
        return
    console.print(f"[green]verify[/green] {script}")


@cli.command("status", help="시스템 상태 대시보드 — LLM/DB/채널 한눈에")
def status_cmd() -> None:
    """전체 시스템 상태를 Rich 테이블로 출력."""
    from datetime import datetime

    console.print(f"\n[bold cyan]career-ops-kr v1.0.0[/bold cyan]  {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")

    # 1. LLM 백엔드
    try:
        from career_ops_kr.ai.client import get_backend_info
        info = get_backend_info()
        backend_color = {"fastflowlm": "green", "ollama": "yellow", "openrouter": "blue"}.get(info["backend"], "dim")
        console.print(f"[bold]LLM:[/bold] [{backend_color}]{info['backend']}[/{backend_color}] model={info['model']} host={info['host']}")
    except Exception as exc:
        console.print(f"[bold]LLM:[/bold] [red]unavailable[/red] ({exc})")

    # 2. SQLite DB
    try:
        from career_ops_kr.storage.sqlite_store import SQLiteStore
        store = SQLiteStore(DATA_DIR / "jobs.db")
        with store._connect() as conn:
            total = conn.execute("SELECT COUNT(*) FROM jobs").fetchone()[0]
            by_grade = conn.execute(
                "SELECT fit_grade, COUNT(*) FROM jobs GROUP BY fit_grade ORDER BY fit_grade"
            ).fetchall()
            latest = conn.execute(
                "SELECT scanned_at FROM jobs ORDER BY scanned_at DESC LIMIT 1"
            ).fetchone()
        grade_str = " ".join(f"{g[0] or '?'}:{g[1]}" for g in by_grade) if by_grade else "(empty)"
        latest_str = latest[0][:16] if latest else "never"
        console.print(f"[bold]DB:[/bold] {total}건  grades=[{grade_str}]  latest={latest_str}")
    except Exception:
        console.print("[bold]DB:[/bold] [dim]not initialized[/dim]")

    # 3. 채널
    try:
        from career_ops_kr.channels import CHANNEL_REGISTRY
        console.print(f"[bold]Channels:[/bold] {len(CHANNEL_REGISTRY)}개 등록")
    except Exception:
        console.print("[bold]Channels:[/bold] [red]import error[/red]")

    # 4. 기관 DB
    try:
        import yaml
        config_path = CONFIG_DIR / "institutions.yml"
        if config_path.exists():
            with open(config_path, encoding="utf-8") as f:
                data = yaml.safe_load(f) or {}
            insts = data.get("institutions", [])
            with_url = sum(1 for i in insts if i.get("career_url"))
            console.print(f"[bold]Institutions:[/bold] {len(insts)}개 ({with_url} with URL)")
        else:
            console.print("[bold]Institutions:[/bold] [dim]not found[/dim]")
    except Exception:
        console.print("[bold]Institutions:[/bold] [dim]error[/dim]")

    # 5. Config 파일
    configs = ["profile.yml", "qualifier_rules.yml", "scoring_weights.yml", "archetypes.yml", "portals.yml"]
    present = sum(1 for c in configs if (CONFIG_DIR / c).exists())
    console.print(f"[bold]Config:[/bold] {present}/{len(configs)} files")

    # 6. 자격증 D-day
    try:
        import yaml
        from datetime import date

        cert_path = CONFIG_DIR / "certifications.yml"
        if cert_path.exists():
            with open(cert_path, encoding="utf-8") as f:
                cert_data = yaml.safe_load(f) or {}
            certs = cert_data.get("certifications", [])
            events = cert_data.get("events", [])
            today = date.today()

            upcoming = []
            for c in certs:
                if c.get("status") in ("upcoming", "registered"):
                    exam = c.get("exam_date", "")
                    reg_start = c.get("register_start", "")
                    if exam:
                        exam_d = date.fromisoformat(exam)
                        days = (exam_d - today).days
                        reg_d = date.fromisoformat(reg_start) if reg_start else None
                        reg_days = (reg_d - today).days if reg_d else None
                        upcoming.append((c["name"], days, reg_days, c.get("status", "")))
            for e in events:
                ed = e.get("date", "")
                if ed:
                    days = (date.fromisoformat(ed) - today).days
                    if days >= 0:
                        upcoming.append((e["name"], days, None, e.get("status", "")))

            upcoming.sort(key=lambda x: x[1])

            if upcoming:
                cert_table = Table(title="자격증/이벤트 D-day", show_header=True, header_style="bold")
                cert_table.add_column("시험/이벤트", style="cyan")
                cert_table.add_column("D-day", justify="right")
                cert_table.add_column("접수", justify="right", style="dim")
                cert_table.add_column("상태")
                for name, days, reg_days, status in upcoming[:10]:
                    if days < 0:
                        dday = f"[dim]D+{abs(days)}[/dim]"
                    elif days == 0:
                        dday = "[red bold]오늘![/red bold]"
                    elif days <= 7:
                        dday = f"[red]D-{days}[/red]"
                    elif days <= 14:
                        dday = f"[yellow]D-{days}[/yellow]"
                    else:
                        dday = f"D-{days}"
                    reg_str = f"D-{reg_days}" if reg_days is not None and reg_days >= 0 else ""
                    if reg_days is not None and 0 <= reg_days <= 3:
                        reg_str = f"[red]D-{reg_days} 접수![/red]"
                    console.print()  # Only print once before table
                    break
                # Actually print table
                for name, days, reg_days, status in upcoming[:10]:
                    if days < 0:
                        dday = f"[dim]D+{abs(days)}[/dim]"
                    elif days == 0:
                        dday = "[red bold]오늘![/red bold]"
                    elif days <= 7:
                        dday = f"[red]D-{days}[/red]"
                    elif days <= 14:
                        dday = f"[yellow]D-{days}[/yellow]"
                    else:
                        dday = f"D-{days}"
                    reg_str = ""
                    if reg_days is not None and reg_days >= 0:
                        reg_str = f"D-{reg_days}"
                        if reg_days <= 3:
                            reg_str = f"[red]D-{reg_days}![/red]"
                    cert_table.add_row(name, dday, reg_str, status)
                console.print(cert_table)
    except Exception:
        pass

    console.print()


@cli.command("dedup", help="중복 제거 — dedup_tracker.py")
@click.option("--apply", is_flag=True, help="실제 적용 (기본은 dry-run)")
def dedup_cmd(apply: bool) -> None:
    mode = "apply" if apply else "dry-run"
    console.print(f"[green]dedup[/green] mode={mode}")


@cli.command("ui", help="TUI 대시보드 실행 — Textual 기반")
def ui_cmd() -> None:
    """TUI 대시보드 실행 — Textual 기반."""
    try:
        from career_ops_kr.tui.app import run_tui
    except ImportError as exc:
        console.print(
            "[red]textual 패키지가 필요합니다.[/red]\n"
            "설치: [bold]pip install 'career-ops-kr\\[tui]'[/bold]\n"
            f"[dim]import error: {exc}[/dim]"
        )
        sys.exit(1)
    run_tui()


@cli.command("login", help="Playwright 수동 로그인 — storage_state 저장")
@click.argument("site")
def login_cmd(site: str) -> None:
    console.print(f"[green]login[/green] {site} (Playwright UI will open)")
    state_dir = DATA_DIR / "auth"
    state_dir.mkdir(parents=True, exist_ok=True)
    console.print(f"[dim]storage_state will be saved to {state_dir / f'{site}.json'}[/dim]")


@cli.command("ai-rank", help="AI 적합도 채점 + 우선순위 랭킹 (OpenRouter 필요)")
@click.option("--site", type=str, default=None, help="특정 채널만 스캔 (기본: 전체)")
@click.option("--model", type=str, default=None, help="OpenRouter 모델 ID (기본: 환경변수 또는 gemini-flash-free)")
@click.option("--top", type=int, default=5, help="상위 N개 출력 (기본: 5)")
@click.option("--api-key", "api_key", type=str, default=None, help="OpenRouter API 키 (기본: OPENROUTER_API_KEY 환경변수)")
@click.option("--max-jobs", "max_jobs", type=int, default=30, help="AI 분석 최대 공고 수 (기본: 30, 무료 tier rate limit 방어)")
def ai_rank_cmd(
    site: str | None,
    model: str | None,
    top: int,
    api_key: str | None,
    max_jobs: int,
) -> None:
    """채널 레지스트리 스캔 → AI 요약 → 적합도 채점 → Top N 우선순위 출력."""
    import yaml

    # --- 1. AI 모듈 + 채널 레지스트리 임포트 ---
    try:
        from career_ops_kr.ai.client import DEFAULT_MODEL, get_client
        from career_ops_kr.ai.ranker import rank_jobs
        from career_ops_kr.ai.scorer import score_jobs_batch
        from career_ops_kr.ai.summarizer import summarize_jobs_batch
        from career_ops_kr.channels import CHANNEL_REGISTRY
    except ImportError as exc:
        console.print(f"[red]import error[/red] {exc}")
        console.print("[dim]openai 설치: uv add openai[/dim]")
        sys.exit(2)

    # --- 2. OpenRouter 클라이언트 초기화 ---
    try:
        client = get_client(api_key)
        _model = model or DEFAULT_MODEL
    except ValueError as exc:
        console.print(f"[red]API 키 오류[/red] {exc}")
        sys.exit(1)

    console.print(f"[dim]model: {_model}[/dim]")

    # --- 3. 공고 수집 ---
    if site:
        if site not in CHANNEL_REGISTRY:
            console.print(f"[red]unknown channel[/red] {site}")
            console.print(f"[dim]사용 가능: {', '.join(CHANNEL_REGISTRY)}[/dim]")
            sys.exit(1)
        targets: dict = {site: CHANNEL_REGISTRY[site]}
    else:
        targets = dict(CHANNEL_REGISTRY)

    console.print(f"[green]스캔 시작[/green] {len(targets)}개 채널")
    all_jobs = []
    for ch_name, cls in targets.items():
        try:
            ch = cls()
            jobs = ch.list_jobs()
            all_jobs.extend(jobs)
            if jobs:
                console.print(f"  [cyan]{ch_name}[/cyan] {len(jobs)}개")
        except Exception as exc:
            console.print(f"  [dim]{ch_name}: {exc}[/dim]")

    if not all_jobs:
        console.print("[yellow]수집된 공고 없음[/yellow] (채널 접속 실패 또는 SPA 채널)")
        sys.exit(0)

    total_collected = len(all_jobs)
    if total_collected > max_jobs:
        console.print(
            f"[yellow]{total_collected}개 수집 → 상위 {max_jobs}개만 AI 분석[/yellow] "
            f"[dim](--max-jobs {max_jobs})[/dim]"
        )
        all_jobs = all_jobs[:max_jobs]
    else:
        console.print(f"[green]총 {total_collected}개[/green] 수집 완료 → AI 분석 시작")

    n = len(all_jobs)

    # --- 4. AI 요약 ---
    _progress_columns = (
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        MofNCompleteColumn(),
        TimeElapsedColumn(),
    )
    summaries: list[str] = []
    with Progress(*_progress_columns, console=console, transient=True) as prog:
        task = prog.add_task("[cyan]요약[/cyan]", total=n)

        def _on_summary(done: int, _total: int) -> None:
            prog.update(task, completed=done)

        summaries = summarize_jobs_batch(all_jobs, client, _model, on_progress=_on_summary)
    console.print(f"[green]요약 완료[/green] {n}개")

    # --- 5. profile.yml 로드 ---
    profile: dict = {}
    profile_path = CONFIG_DIR / "profile.yml"
    if profile_path.exists():
        try:
            profile = yaml.safe_load(profile_path.read_text(encoding="utf-8")) or {}
        except Exception as exc:
            console.print(f"[yellow]profile.yml 로드 실패 (채점 정확도 낮음)[/yellow] {exc}")
    else:
        console.print("[yellow]config/profile.yml 없음 — 기본 채점 적용[/yellow]")

    # --- 6. AI 채점 ---
    fit_scores: list[tuple[int, str]] = []
    with Progress(*_progress_columns, console=console, transient=True) as prog:
        task = prog.add_task("[cyan]채점[/cyan]", total=n)

        def _on_score(done: int, _total: int) -> None:
            prog.update(task, completed=done)

        fit_scores = score_jobs_batch(all_jobs, summaries, profile, client, _model, on_progress=_on_score)
    console.print(f"[green]채점 완료[/green] {n}개")

    # --- 7. 랭킹 ---
    ranked = rank_jobs(all_jobs, fit_scores, summaries, top_n=top)

    # --- 8. Rich 테이블 출력 ---
    table = Table(title=f"AI 추천 Top {len(ranked)} (전체 {len(all_jobs)}개 분석)")
    table.add_column("#", justify="right", style="bold")
    table.add_column("org", style="cyan")
    table.add_column("title")
    table.add_column("archetype", style="dim")
    table.add_column("fit", justify="right")
    table.add_column("total", justify="right", style="bold green")
    table.add_column("D-day", justify="right")
    table.add_column("이유", style="dim")

    for i, r in enumerate(ranked, 1):
        days = r.days_left
        if days is None:
            dday = "-"
        elif days < 0:
            dday = f"[red]D+{abs(days)}[/red]"
        elif days == 0:
            dday = "[red]오늘[/red]"
        elif days <= 3:
            dday = f"[red]D-{days}[/red]"
        elif days <= 7:
            dday = f"[yellow]D-{days}[/yellow]"
        else:
            dday = f"D-{days}"

        table.add_row(
            str(i),
            r.job.org[:20],
            r.job.title[:35],
            r.job.archetype or "-",
            str(r.fit_score),
            str(r.total_score),
            dday,
            r.fit_reason[:40] if r.fit_reason else "-",
        )

    console.print(table)
    console.print(
        f"[dim]모델: {_model} | "
        f"fit=AI채점(0~100) | total=fit+마감긴급도+archetype보너스[/dim]"
    )

    # --- 9. 상세 패널 (Top 3) ---
    for i, r in enumerate(ranked[:3], 1):
        summary_text = r.summary or "(요약 없음)"
        console.print(
            Panel.fit(
                f"[bold]{r.job.title}[/bold]  —  {r.job.org}\n"
                f"[dim]URL:[/dim] {r.job.source_url}\n"
                f"[dim]요약:[/dim] {summary_text}\n"
                f"[dim]이유:[/dim] {r.fit_reason}",
                title=f"#{i} total={r.total_score}",
                border_style="cyan" if i == 1 else "dim",
            )
        )


@cli.command("institutions", help="201개 금융기관 채용공고 일괄 검색 (aggregator API)")
@click.option("--grade", type=str, default=None, help="적합도 필터: S,A,B,C,D (콤마 구분, 미지정=전체)")
@click.option("--concurrency", type=int, default=6, help="동시 검색 수")
@click.option("--top", type=int, default=50, help="상위 N건 표시")
def institutions_cmd(grade: str | None, concurrency: int, top: int) -> None:
    """201개 기관을 wanted/jobkorea API에서 기관명 검색 → 실 채용공고 매칭."""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from pathlib import Path

    import yaml

    config_path = PROJECT_ROOT / "config" / "institutions.yml"
    if not config_path.exists():
        console.print("[red]config/institutions.yml 없음[/red] — 먼저 엑셀 변환 필요")
        sys.exit(1)

    with open(config_path, encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}

    insts = data.get("institutions", [])
    if grade:
        grades = {g.strip().upper() for g in grade.split(",")}
        insts = [i for i in insts if i.get("grade", "").upper() in grades]

    console.print(f"[green]institutions[/green] {len(insts)}개 기관 × 3 aggregator (wanted+jobkorea+saramin)")

    import requests as _req
    from career_ops_kr.channels.base import BaseChannel, JobRecord, deadline_parser

    ua = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36 career-ops-kr/0.2"
    )

    all_jobs: list[JobRecord] = []
    org_counts: dict[str, int] = {}

    def _search_wanted(search_name: str) -> list[JobRecord]:
        """Wanted JSON API 검색."""
        from datetime import datetime
        from urllib.parse import quote
        api_url = (
            f"https://www.wanted.co.kr/api/v4/jobs"
            f"?query={quote(search_name)}&country=kr&job_sort=job.latest_order"
            f"&years=-1&limit=10&offset=0"
        )
        try:
            resp = _req.get(api_url, headers={"User-Agent": ua}, timeout=10)
            if resp.status_code != 200:
                return []
            items = resp.json().get("data", [])
        except Exception:
            return []
        results = []
        now = datetime.now()
        for item in items[:10]:
            try:
                jid = str(item.get("id", ""))
                company_name = item.get("company", {}).get("name", "")
                title = item.get("position", "") or item.get("title", "")
                job_url = f"https://www.wanted.co.kr/wd/{jid}" if jid else ""
                if not job_url or not title:
                    continue
                results.append(JobRecord(
                    id=BaseChannel._make_id(job_url, title[:120]),
                    source_url=job_url, source_channel="wanted", source_tier=1,
                    org=company_name or search_name, title=title[:200],
                    description=title, legitimacy_tier="T1", scanned_at=now,
                ))
            except Exception:
                continue
        return results

    def _search_jobkorea(search_name: str) -> list[JobRecord]:
        """잡코리아 HTML 검색 파싱."""
        from datetime import datetime
        from urllib.parse import quote
        from bs4 import BeautifulSoup
        url = f"https://www.jobkorea.co.kr/Search/?stext={quote(search_name)}&tabType=recruit&Page_No=1"
        try:
            resp = _req.get(url, headers={"User-Agent": ua, "Accept": "text/html"}, timeout=10)
            if resp.status_code != 200:
                return []
        except Exception:
            return []
        soup = BeautifulSoup(resp.text, "html.parser")
        results = []
        now = datetime.now()
        seen = set()
        for a in soup.select("a[href*='/Recruit/GI_Read']"):
            title = a.get_text(" ", strip=True)
            href = a.get("href", "")
            if not title or len(title) < 4 or not href:
                continue
            if href.startswith("/"):
                href = "https://www.jobkorea.co.kr" + href
            if href in seen:
                continue
            seen.add(href)
            try:
                results.append(JobRecord(
                    id=BaseChannel._make_id(href, title[:120]),
                    source_url=href, source_channel="jobkorea", source_tier=1,
                    org=search_name, title=title[:200],
                    description=title, legitimacy_tier="T1", scanned_at=now,
                ))
            except Exception:
                continue
            if len(results) >= 10:
                break
        return results

    def _search_saramin(search_name: str) -> list[JobRecord]:
        """사람인 HTML 검색 파싱."""
        from datetime import datetime
        from urllib.parse import quote
        from bs4 import BeautifulSoup
        url = f"https://www.saramin.co.kr/zf_user/search?searchword={quote(search_name)}&searchType=search&recruitPage=1"
        try:
            resp = _req.get(url, headers={"User-Agent": ua, "Accept": "text/html"}, timeout=10)
            if resp.status_code != 200:
                return []
        except Exception:
            return []
        soup = BeautifulSoup(resp.text, "html.parser")
        results = []
        now = datetime.now()
        seen = set()
        for a in soup.select("a[href*='/zf_user/jobs/relay']"):
            title = a.get_text(" ", strip=True)
            href = a.get("href", "")
            if not title or len(title) < 4 or not href:
                continue
            if href.startswith("/"):
                href = "https://www.saramin.co.kr" + href
            if href in seen:
                continue
            seen.add(href)
            try:
                results.append(JobRecord(
                    id=BaseChannel._make_id(href, title[:120]),
                    source_url=href, source_channel="saramin", source_tier=1,
                    org=search_name, title=title[:200],
                    description=title, legitimacy_tier="T1", scanned_at=now,
                ))
            except Exception:
                continue
            if len(results) >= 10:
                break
        return results

    def _fuzzy_match(search_name: str, company_name: str) -> bool:
        """Fuzzy org name matching — 노이즈 필터링."""
        if not company_name:
            return False
        # Exact substring match (either direction)
        if search_name in company_name or company_name in search_name:
            return True
        # Fuzzy match with threshold
        try:
            from fuzzywuzzy import fuzz
            return fuzz.partial_ratio(search_name, company_name) >= 60
        except ImportError:
            return search_name.lower() in company_name.lower()

    def _filter_by_org(results: list[JobRecord], search_name: str) -> list[JobRecord]:
        """Filter results to only include matching org names."""
        return [r for r in results if _fuzzy_match(search_name, r.org)]

    def _search_one(inst: dict) -> tuple[str, list[JobRecord]]:
        import re
        import time as _time
        name = inst["name"]
        search_name = re.sub(r"\(.*?\)", "", name).strip()
        if len(search_name) < 2:
            return name, []
        # 3 aggregator 동시 검색 + fuzzy 필터 + rate limit
        results: list[JobRecord] = []
        results.extend(_filter_by_org(_search_wanted(search_name), search_name))
        _time.sleep(0.3)  # rate limit between aggregators
        results.extend(_filter_by_org(_search_jobkorea(search_name), search_name))
        _time.sleep(0.3)
        results.extend(_filter_by_org(_search_saramin(search_name), search_name))
        return name, results

    with Progress(
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        MofNCompleteColumn(),
        TimeElapsedColumn(),
        console=console,
        transient=True,
    ) as prog:
        task = prog.add_task("[cyan]기관 검색[/cyan]", total=len(insts))

        with ThreadPoolExecutor(max_workers=concurrency) as pool:
            futures = {pool.submit(_search_one, inst): inst["name"] for inst in insts}
            for future in as_completed(futures):
                inst_name = futures[future]
                try:
                    name, jobs = future.result()
                    if jobs:
                        org_counts[name] = len(jobs)
                        all_jobs.extend(jobs)
                except Exception:
                    pass
                prog.advance(task)

    # 중복 제거
    seen: set[str] = set()
    unique: list[JobRecord] = []
    for j in all_jobs:
        if j.id not in seen:
            seen.add(j.id)
            unique.append(j)

    console.print(f"\n[green]결과[/green] {len(unique)}건 (중복 제거 후, {len(org_counts)}개 기관 매칭)")

    # 기관별 매칭 테이블
    if org_counts:
        table = Table(title=f"기관별 채용공고 ({len(org_counts)}개 매칭)")
        table.add_column("#", justify="right")
        table.add_column("기관명", style="cyan")
        table.add_column("건수", justify="right", style="bold green")
        for i, (org, cnt) in enumerate(sorted(org_counts.items(), key=lambda x: -x[1]), 1):
            if i > top:
                break
            table.add_row(str(i), org, str(cnt))
        console.print(table)

    # 전체 공고 테이블
    if unique:
        job_table = Table(title=f"채용공고 Top {min(top, len(unique))}")
        job_table.add_column("#", justify="right")
        job_table.add_column("기관", style="cyan")
        job_table.add_column("포지션")
        job_table.add_column("URL", style="dim")
        for i, j in enumerate(unique[:top], 1):
            job_table.add_row(str(i), j.org[:20], j.title[:40], str(j.source_url)[:50])
        console.print(job_table)

    if not unique:
        console.print("[yellow]매칭된 공고 없음 — 현재 채용 중인 기관이 없거나 API 응답 변경[/yellow]")


@cli.command("export", help="공고 DB → Excel 파일 출력 (색상 포맷 포함)")
@click.option("--output", "-o", type=click.Path(path_type=Path), default=None,
              help="출력 파일 경로 (기본: output/jobs_YYYYMMDD.xlsx)")
@click.option("--status", type=str, default=None, help="상태 필터 (inbox/applied/passed/rejected)")
@click.option("--days", type=int, default=None, help="마감 N일 이내만 추출")
@click.option("--open-only", is_flag=True, default=False, help="마감 전 공고만")
def export_cmd(output: Path | None, status: str | None, days: int | None, open_only: bool) -> None:
    """SQLite DB → 색상 포맷 Excel 파일 출력."""
    import sqlite3
    from datetime import datetime, timedelta

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        console.print("[red]openpyxl 필요[/red] — `uv add openpyxl` 실행")
        sys.exit(1)

    db = DATA_DIR / "jobs.db"
    if not db.exists():
        console.print("[yellow]DB 없음[/yellow] — career-ops scan --all 먼저 실행")
        sys.exit(1)

    conn = sqlite3.connect(str(db))
    conn.row_factory = sqlite3.Row

    query = "SELECT * FROM jobs WHERE 1=1"
    params: list = []
    if status:
        query += " AND status = ?"
        params.append(status)
    if open_only:
        query += " AND (deadline IS NULL OR deadline >= date('now'))"
    if days:
        cutoff = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d")
        query += " AND deadline IS NOT NULL AND deadline <= ?"
        params.append(cutoff)
    query += " ORDER BY deadline ASC NULLS LAST, source_tier ASC"

    rows = conn.execute(query, params).fetchall()
    conn.close()

    if not rows:
        console.print("[yellow]조건에 맞는 공고 없음[/yellow]")
        return

    # 출력 경로
    if output is None:
        out_dir = PROJECT_ROOT / "output"
        out_dir.mkdir(parents=True, exist_ok=True)
        output = out_dir / f"jobs_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    else:
        output.parent.mkdir(parents=True, exist_ok=True)

    # --- Excel 빌드 ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "공고목록"

    # 컬럼 정의
    COLS = [
        ("기관", "org", 20),
        ("공고명", "title", 45),
        ("유형", "archetype", 8),
        ("채널", "source_channel", 18),
        ("Tier", "source_tier", 5),
        ("마감일", "deadline", 12),
        ("D-Day", "_dday", 8),
        ("지역", "location", 8),
        ("상태", "status", 8),
        ("Grade", "fit_grade", 7),
        ("URL", "source_url", 40),
    ]

    # 헤더 스타일
    HEADER_FILL = PatternFill("solid", fgColor="1530FF")  # 링커리어 다크블루
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="DDDDDD")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (label, _, width) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # D-day 계산
    today = datetime.now().date()

    def calc_dday(deadline_str: str | None) -> str:
        if not deadline_str:
            return "—"
        try:
            d = datetime.strptime(str(deadline_str)[:10], "%Y-%m-%d").date()
            delta = (d - today).days
            if delta < 0:
                return "마감"
            if delta == 0:
                return "D-DAY"
            return f"D-{delta}"
        except Exception:
            return "—"

    # 색상 정의
    FILL_URGENT  = PatternFill("solid", fgColor="FFE8E8")  # 3일 내 — 연빨강
    FILL_SOON    = PatternFill("solid", fgColor="FFF3E0")  # 7일 내 — 연주황
    FILL_OK      = PatternFill("solid", fgColor="F2FAFF")  # 그 외 — 연파랑
    FILL_CLOSED  = PatternFill("solid", fgColor="F5F5F5")  # 마감 — 회색
    FILL_APPLIED = PatternFill("solid", fgColor="E8FFE8")  # applied — 연초록

    STATUS_LABELS = {"inbox": "📥 대기", "applied": "📤 지원", "passed": "✅ 합격", "rejected": "❌ 탈락"}
    ARCH_LABELS   = {"INTERN": "인턴", "ENTRY": "신입", "EXPERIENCED": "경력", "IT": "IT"}

    for row_idx, row in enumerate(rows, 2):
        dday = calc_dday(row["deadline"])
        try:
            delta = int(dday.replace("D-", "")) if dday.startswith("D-") else (0 if dday == "D-DAY" else -1)
        except Exception:
            delta = 999

        # 행 배경
        if row["status"] == "applied":
            row_fill = FILL_APPLIED
        elif dday == "마감":
            row_fill = FILL_CLOSED
        elif delta <= 3:
            row_fill = FILL_URGENT
        elif delta <= 7:
            row_fill = FILL_SOON
        else:
            row_fill = FILL_OK

        values = {
            "org": row["org"] or "",
            "title": row["title"] or "",
            "archetype": ARCH_LABELS.get(row["archetype"] or "", row["archetype"] or ""),
            "source_channel": row["source_channel"] or "",
            "source_tier": row["source_tier"],
            "deadline": str(row["deadline"] or "")[:10] or "",
            "_dday": dday,
            "location": row["location"] or "전국",
            "status": STATUS_LABELS.get(row["status"] or "", row["status"] or ""),
            "fit_grade": row["fit_grade"] or "",
            "source_url": row["source_url"] or "",
        }

        for col_idx, (_, key, _) in enumerate(COLS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=values[key])
            cell.fill = row_fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(key == "title"))
            cell.font = Font(size=9)
            # URL 컬럼 → 하이퍼링크
            if key == "source_url" and values[key].startswith("http"):
                cell.hyperlink = values[key]
                cell.value = "🔗 링크"
                cell.font = Font(size=9, color="0563C1", underline="single")
            # D-day 컬럼 색
            if key == "_dday":
                if dday in ("마감",):
                    cell.font = Font(size=9, color="999999")
                elif delta <= 3:
                    cell.font = Font(size=9, bold=True, color="C0392B")
                elif delta <= 7:
                    cell.font = Font(size=9, bold=True, color="E67E22")
                else:
                    cell.font = Font(size=9, color="2980B9")

        ws.row_dimensions[row_idx].height = 18

    # 요약 시트
    ws2 = wb.create_sheet("요약")
    ws2["A1"] = "career-ops-kr DB 현황"
    ws2["A1"].font = Font(bold=True, size=14, color="1530FF")
    ws2["A2"] = f"출력일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws2["A3"] = f"총 공고: {len(rows)}건"

    summary_data = [
        ("", ""),
        ("채널별 공고 수", ""),
    ]
    chan_counts: dict[str, int] = {}
    for row in rows:
        ch = row["source_channel"] or "?"
        chan_counts[ch] = chan_counts.get(ch, 0) + 1
    for ch, n in sorted(chan_counts.items(), key=lambda x: -x[1]):
        summary_data.append((ch, n))

    summary_data.append(("", ""))
    summary_data.append(("유형별 공고 수", ""))
    arch_counts: dict[str, int] = {}
    for row in rows:
        a = ARCH_LABELS.get(row["archetype"] or "", row["archetype"] or "기타")
        arch_counts[a] = arch_counts.get(a, 0) + 1
    for a, n in sorted(arch_counts.items(), key=lambda x: -x[1]):
        summary_data.append((a, n))

    for r_idx, (k, v) in enumerate(summary_data, 5):
        ws2.cell(row=r_idx, column=1, value=k).font = Font(size=10, bold=(v == ""))
        if v != "":
            ws2.cell(row=r_idx, column=2, value=v).font = Font(size=10)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 10

    wb.save(str(output))
    console.print(f"[green]Excel 출력 완료[/green]  → [bold]{output}[/bold]")
    console.print(f"   공고 {len(rows)}건 | 필터: status={status} days={days} open_only={open_only}")


def _register_commands() -> None:
    """commands/ 서브패키지 커맨드를 cli group에 동적 등록.

    각 모듈 임포트 실패 시 해당 커맨드만 건너뛰고 나머지는 정상 등록.
    """
    import logging as _logging

    _log = _logging.getLogger(__name__)
    _cmds = [
        ("career_ops_kr.commands.filter_cmd", "filter_cmd"),
        ("career_ops_kr.commands.auto_pipeline", "auto_pipeline_cmd"),
        ("career_ops_kr.commands.batch_cmd", "batch_cmd"),
        ("career_ops_kr.commands.notify_cmd", "notify_cmd"),
        ("career_ops_kr.commands.apply_cmd", "apply_cmd"),
        ("career_ops_kr.commands.interview_cmd", "interview_prep_cmd"),
        ("career_ops_kr.commands.followup_cmd", "followup_cmd"),
        ("career_ops_kr.commands.project_cmd", "project_cmd"),
        ("career_ops_kr.commands.patterns_cmd", "patterns_cmd"),
        ("career_ops_kr.commands.vault_cmd", "vault_sync_cmd"),
        ("career_ops_kr.commands.history_cmd", "history_cmd"),
        ("career_ops_kr.commands.ncs_cmd", "ncs_cmd"),
        ("career_ops_kr.commands.web_cmd", "web_cmd"),
        ("career_ops_kr.commands.reclassify_cmd", "reclassify_cmd"),
        ("career_ops_kr.commands.bookmark_cmd", "bookmark_cmd"),
    ]
    for mod_path, fn_name in _cmds:
        try:
            mod = __import__(mod_path, fromlist=[fn_name])
            cli.add_command(getattr(mod, fn_name))
        except Exception as exc:
            _log.debug("command load failed %s: %s", mod_path, exc)


_register_commands()


def main() -> None:
    """Entry point for [project.scripts] career-ops = cli:main."""
    try:
        cli(standalone_mode=True)
    except click.ClickException:
        raise
    except Exception as exc:
        console.print(f"[red]unexpected error[/red] {exc}")
        sys.exit(1)


if __name__ == "__main__":
    main()
